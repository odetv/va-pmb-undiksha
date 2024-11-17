import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import random
import string

def generate_id():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def log_activity(log):
    log_dir = "api/logs"
    os.makedirs(log_dir, exist_ok=True)
    file_path = os.path.join(log_dir, "log_activity.xlsx")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Timestamp", "Method", "Status Code", "Success", "Description"])
    ws.append([
        log["id"],
        log["timestamp"],
        log["method"],
        log["status_code"],
        log["success"],
        log["description"]
    ])
    wb.save(file_path)


def log_configllm(log):
    log_dir = "api/logs"
    os.makedirs(log_dir, exist_ok=True)
    file_path = os.path.join(log_dir, "log_configllm.xlsx")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "LLM", "Model LLM", "Embedder", "Model Embedder", "Chunk Size", "Chunk Overlap", "Total Chunks"])
    ws.delete_rows(2, ws.max_row)
    ws.append([
        log["timestamp"],
        log["llm"],
        log["model_llm"],
        log["embbeder"],
        log["model_embedder"],
        log["chunk_size"],
        log["chunk_overlap"],
        log["total_chunks"]
    ])
    wb.save(file_path)
