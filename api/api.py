import json
import os
import shutil
from datetime import datetime
from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.exceptions import RequestValidationError
from fastapi.security import APIKeyHeader
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.status import HTTP_404_NOT_FOUND, HTTP_405_METHOD_NOT_ALLOWED
from dotenv import load_dotenv
from pydantic import BaseModel, Field, ValidationError
from typing_extensions import List
from langchain_community.document_loaders import PyPDFDirectoryLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain_openai import OpenAIEmbeddings
from langchain_community.embeddings.ollama import OllamaEmbeddings
from main import build_graph
from utils.scrapper_rss import scrap_news
from utils.logging import generate_id, log_activity, log_configllm
from openpyxl import load_workbook
from src.config.config import DATASETS_DIR, VECTORDB_DIR


load_dotenv()
bearer_token = os.getenv("VA_BEARER_TOKEN")
openai_api_key = os.getenv("OPENAI_API_KEY")
ollama_base_url = os.getenv("OLLAMA_BASE_URL")
bearer_token_header = APIKeyHeader(name="Authorization", auto_error=False)


#  Autherization bearer token API
def verify_bearer_token(request_http: Request, token: str = Depends(bearer_token_header)):
    if token is None or not token.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Eitss... mau ngapain? Akses terbatas!")
    actual_token = token[7:]
    if actual_token != bearer_token:
        raise HTTPException(status_code=401, detail="Invalid Token")


# Class API model
class QuestionRequest(BaseModel):
    question: str
class QuestionResponse(BaseModel):
    question: str
    answer: str
class DeleteDatasetsRequest(BaseModel):
    filenames: List[str]
class ProcessRequest(BaseModel):
    llm: str
    model_llm: str
    embbeder: str
    model_embedder: str
    chunk_size: int
    chunk_overlap: int
    class Config:
        protected_namespaces = ()


# Metadata API
tags_metadata = [
    {
        "name": "root",
        "description": "Status kondisi service API dengan timestamp."
    },
    {
        "name": "datasets",
        "description": "Proses menyiapkan datasets (CRUD)."
    },
    {
        "name": "setup",
        "description": "Proses load dokumen, chunking, embedding, dan membuat vector database."
    },
    {
        "name": "checkmodel",
        "description": "Cek parameter model yang digunakan."
    },
    {
        "name": "news",
        "description": "Hasil scrapping news."
    },
    {
        "name": "graph",
        "description": "Gambar alur graph terbentuk."
    },
    {
        "name": "chat",
        "description": "Percakapan pengguna dengan virtual assistant."
    },
]


# Initialize FastAPI
app = FastAPI(
    openapi_tags=tags_metadata,
    title="API PMB Undiksha",
    summary="API PMB Undiksha",
    version="0.0.1",
    docs_url="/docs",
    redoc_url=None,
    openapi_url="/openapipmb.json"
)


# CORS Headers
origins = ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Format API response
def api_response(status_code: int, success: bool, message: str, data=None):
    return JSONResponse(
        status_code=status_code,
        content={
            "statusCode": status_code,
            "success": success,
            "message": message,
            "data": data
        }
    )


# Enpoint untuk base url root API request
@app.get("/", tags=["root"])
async def root(request_http: Request, token: str = Depends(verify_bearer_token)):
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request_http.method} {request_http.url.path}",
        "status_code": 200,
        "success": True,
        "description": "API Virtual Assistant Undiksha "+datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    return api_response(
        status_code=200,
        success=True,
        message="OK",
        data={"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "description": "API Virtual Assistant Undiksha"}
    )


# Endpoint untuk melihat daftar file (List)
@app.get("/datasets/list", tags=["datasets"])
async def list_datasets(request_http: Request, token: str = Depends(verify_bearer_token)):
    if not os.path.exists(DATASETS_DIR):
        os.makedirs(DATASETS_DIR)
        raise HTTPException(status_code=404, detail="Folder datasets dibuat, tetapi masih kosong.")
    datasets = [
        file for file in os.listdir(DATASETS_DIR)
        if file.lower().endswith(('.pdf', '.doc', '.docx'))
    ]
    if not datasets:
        raise HTTPException(status_code=404, detail="Folder datasets kosong atau tidak ada file PDF/Word.")
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request_http.method} {request_http.url.path}",
        "status_code": 200,
        "success": True,
        "description": f"Datasets ditemukan.\n###\nDatasets:{datasets}"
    })
    return api_response(
        status_code=200,
        success=True,
        message="Datasets ditemukan.",
        data=datasets
    )


# Endpoint untuk membaca konten file tertentu (Read)
@app.get("/datasets/read/{filename}", tags=["datasets"])
async def read_datasets(request_http: Request, filename: str, token: str = Depends(verify_bearer_token)):
    file_path = os.path.join(DATASETS_DIR, filename)
    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File tidak ditemukan.")
    if filename.lower().endswith(".pdf"):
        def iter_file():
            with open(file_path, "rb") as file:
                yield from file
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"application/pdf\n###\nfilename={filename}"
        })
        return StreamingResponse(iter_file(), media_type="application/pdf", headers={"Content-Disposition": f"inline; filename={filename}"})
    elif filename.lower().endswith((".doc", ".docx")):
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"application/vnd.openxmlformats-officedocument.wordprocessingml.document\n###\nfilename={filename}"
        })
        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", headers={"Content-Disposition": f"attachment; filename={filename}"})
    else:
        raise HTTPException(status_code=400, detail="Format file tidak didukung untuk ditampilkan atau diunduh.")


# Endpoint untuk mengunggah file PDF/Word single atau multiple files (Create)
@app.post("/datasets/upload", tags=["datasets"])
async def upload_datasets(request_http: Request, files: List[UploadFile] = File(...), token: str = Depends(verify_bearer_token)):
    uploaded_files = []
    unsupported_files = []
    for file in files:
        if file.content_type not in ["application/pdf", "application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
            unsupported_files.append(file.filename)
            continue
        file_location = os.path.join(DATASETS_DIR, file.filename)
        with open(file_location, "wb") as f:
            shutil.copyfileobj(file.file, f)
        uploaded_files.append(file.filename)
    if not uploaded_files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diunggah atau file yang diunggah tidak didukung.")
    if unsupported_files:
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 207,
            "success": True,
            "description": f"Beberapa file tidak didukung.\n##\nuploaded_files:{uploaded_files}\n##\nunsupported_files:{unsupported_files}"
        })
        return api_response(
            status_code=207,
            success=True,
            message="Beberapa file tidak didukung.",
            data={"uploaded_files": uploaded_files, "unsupported_files": unsupported_files}
        )
    else:
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 201,
            "success": True,
            "description": f"Semua file berhasil diunggah.\n##\nuploaded_files:{uploaded_files}"
        })
        return api_response(
            status_code=201,
            success=True,
            message="Semua file berhasil diunggah.",
            data={"uploaded_files": uploaded_files}
        )


# Endpoint untuk memperbarui (Update) file
@app.put("/datasets/update", tags=["datasets"])
async def update_dataset(
    request_http: Request,
    target: str = Form(...),
    file: UploadFile = File(...),
    token: str = Depends(verify_bearer_token)
):
    old_filename = target
    new_filename = file.filename
    original_file_path = os.path.join(DATASETS_DIR, target)
    if not os.path.isfile(original_file_path):
        raise HTTPException(status_code=404, detail="File tidak ditemukan.")
    new_file_path = os.path.join(DATASETS_DIR, new_filename)
    os.remove(original_file_path)
    with open(new_file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request_http.method} {request_http.url.path}",
        "status_code": 200,
        "success": True,
        "description": f"File berhasil diperbarui.\n###\ntarget:{old_filename}\n###\nold_filename:{old_filename}\n###\nnew_filename:{new_filename}"
    })
    return api_response(
        status_code=200,
        success=True,
        message="File berhasil diperbarui.",
        data={
            "target": old_filename,
            "old_filename": old_filename,
            "new_filename": new_filename
        }
    )


# Endpoint untuk menghapus (Delete) file
@app.delete("/datasets/delete", tags=["datasets"])
async def delete_datasets(request_http: Request, request: DeleteDatasetsRequest, token: str = Depends(verify_bearer_token)):
    deleted_files = []
    not_found_files = []
    for filename in request.filenames:
        file_path = os.path.join(DATASETS_DIR, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
            deleted_files.append(filename)
        else:
            not_found_files.append(filename)
    if not_found_files:
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 207,
            "success": True,
            "description": f"Beberapa file tidak ditemukan.\n###\ndeleted_files:{deleted_files}\n###\nnot_found_files:{not_found_files}"
        })
        return api_response(
            status_code=207,
            success=True,
            message="Beberapa file tidak ditemukan.",
            data={"deleted_files": deleted_files, "not_found_files": not_found_files}
        )
    if not deleted_files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang dipilih.")
    else:
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"Semua file berhasil dihapus.\n###\ndeleted_files:{deleted_files}"
        })
        return api_response(
            status_code=200,
            success=True,
            message="Semua file berhasil dihapus.",
            data={"deleted_files": deleted_files}
        )


# Endpoint setup awal untuk raw process vectordb (load dokumen, chunking, dan embedding)
@app.post("/setup", tags=["setup"])
async def raw_process(request_http: Request, request: ProcessRequest, token: str = Depends(verify_bearer_token)):
    if not request.llm:
        raise HTTPException(status_code=400, detail="LLM harus diisi dengan sesuai.")
    if not request.model_llm:
        raise HTTPException(status_code=400, detail="Model LLM harus diisi dengan sesuai.")
    valid_llm = ["openai", "ollama"]
    if request.llm not in valid_llm:
        raise HTTPException(status_code=400, detail="LLM harus 'openai' atau 'ollama'.")
    valid_model_llm = {
        "openai": ["gpt-4o", "gpt-4o-mini"],
        "ollama": ["gemma2", "llama3.1"]
    }
    if request.model_llm not in valid_model_llm.get(request.llm, []):
        raise HTTPException(status_code=400, detail=f"Model LLM untuk '{request.llm}' harus salah satu dari {valid_model_llm[request.llm]}.")
    if not request.embbeder:
        raise HTTPException(status_code=400, detail="Embbeder harus diisi dengan sesuai.")
    if not request.model_embedder:
        raise HTTPException(status_code=400, detail="Model Embedder harus diisi dengan sesuai.")
    valid_embedder = ["openai", "ollama"]
    if request.embbeder not in valid_embedder:
        raise HTTPException(status_code=400, detail="Embedder harus 'openai' atau 'ollama'.")
    valid_embedder_model = {
        "openai": ["text-embedding-3-large", "text-embedding-3-small"],
        "ollama": ["bge-m3", "mxbai-embed-large"]
    }
    if request.model_embedder not in valid_embedder_model.get(request.embbeder, []):
        raise HTTPException(status_code=400, detail=f"Model Embedder untuk '{request.embbeder}' harus salah satu dari {valid_embedder_model[request.embbeder]}.")
    def get_embbeder():
        if request.embbeder.lower() == "openai":
            return OpenAIEmbeddings(api_key=openai_api_key, model=request.model_embedder)
        elif request.embbeder.lower() == "ollama":
            return OllamaEmbeddings(base_url=ollama_base_url, model=request.model_embedder)
    if request.chunk_size <= 0:
        raise HTTPException(status_code=400, detail="Chunk Size harus diisi nilai lebih dari 0.")
    if request.chunk_overlap <= 0:
        raise HTTPException(status_code=400, detail="Chunk Overlap harus diisi nilai lebih dari 0.")
    EMBEDDER = get_embbeder()
    if not os.path.exists(DATASETS_DIR):
        os.makedirs(DATASETS_DIR)
    if not os.path.exists(VECTORDB_DIR):
        os.makedirs(VECTORDB_DIR)
    try:
        loader = PyPDFDirectoryLoader(DATASETS_DIR)
        documents = loader.load()
        if not documents:
            log_activity({
                "id": generate_id(),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "method": f"{request_http.method} {request_http.url.path}",
                "status_code": 404,
                "success": False,
                "description": "Tidak ada dokumen PDF di direktori datasets."
            })
            return api_response(
                status_code=404,
                success=False,
                message="Tidak ada dokumen PDF di direktori datasets.",
                data=None
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saat memuat dokumen: {str(e)}")
    try:
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=request.chunk_size,
            chunk_overlap=request.chunk_overlap
        )
        chunks = text_splitter.split_documents(documents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saat melakukan chunking: {str(e)}")
    try:
        vectordb = FAISS.from_documents(chunks, EMBEDDER)
        vectordb.save_local(VECTORDB_DIR)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saat menyimpan embeddings: {str(e)}")
    log_configllm({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "llm": request.llm,
        "model_llm": request.model_llm,
        "embbeder": request.embbeder,
        "model_embedder": request.model_embedder,
        "chunk_size": request.chunk_size,
        "chunk_overlap": request.chunk_overlap,
        "total_chunks": len(chunks)
    })
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request_http.method} {request_http.url.path}",
        "status_code": 200,
        "success": True,
        "description": f"Proses penyiapan dokumen berhasil diselesaikan dan embeddings berhasil disimpan pada vector database.\n###\nllm:{request.llm}\n###\nmodel_llm:{request.model_llm}\n###\nembbeder:{request.embbeder}\n###\nmodel_embedder:{request.model_embedder}\n###\nchunk_size:{request.chunk_size}\n###\nchunk_overlap:{request.chunk_overlap}\n###\ntotal_chunks:{len(chunks)}"
    })
    return api_response(
        status_code=200,
        success=True,
        message="Proses penyiapan dokumen berhasil diselesaikan dan embeddings berhasil disimpan pada vector database.",
        data={
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "llm": request.llm,
            "model_llm": request.model_llm,
            "embbeder": request.embbeder,
            "model_embedder": request.model_embedder,
            "chunk_size": request.chunk_size,
            "chunk_overlap": request.chunk_overlap,
            "total_chunks": len(chunks)
        }
    )


# Endpoint untuk mengecek LLM dan EMBEDDER
@app.get("/checkmodel", tags=["checkmodel"])
async def check_model(request_http: Request, token: str = Depends(verify_bearer_token)):
    try:
        file_path = "api/logs/log_configllm.xlsx"
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Tidak ditemukan file log config LLM.")
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        last_row = sheet.max_row
        if last_row < 2:
            raise HTTPException(status_code=404, detail="Tidak ditemukan data log config LLM.")
        data = {
            "last_update": sheet.cell(row=last_row, column=1).value or "Unknown",
            "llm": sheet.cell(row=last_row, column=2).value or "",
            "model_llm": sheet.cell(row=last_row, column=3).value or "",
            "embbeder": sheet.cell(row=last_row, column=4).value or "",
            "model_embedder": sheet.cell(row=last_row, column=5).value or "",
            "chunk_size": sheet.cell(row=last_row, column=6).value or "",
            "chunk_overlap": sheet.cell(row=last_row, column=7).value or "",
            "total_chunks": sheet.cell(row=last_row, column=8).value or ""
        }
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"OK\n###\nLog Config LLM:{data}"
        })
        return api_response(
            status_code=200,
            success=True,
            message="OK",
            data=data
        )
    except HTTPException as e:
        raise HTTPException(status_code=e.status_code, detail=f"{e.detail}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{str(e)}")


# Endpoint untuk scrapping berita
@app.get("/news", tags=["news"])
async def scrapping_news(request_http: Request, token: str = Depends(verify_bearer_token)):
    news_data = scrap_news()
    news_data_json = json.loads(news_data)
    try:
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"OK.\n###\nNews:{news_data_json}"
        })
        return api_response(
                status_code=200,
                success=True,
                message="OK",
                data=news_data_json
            )
    except HTTPException as e:
        raise HTTPException(status_code=e.status_code, detail=f"{e.detail}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{str(e)}")


# Endpoint untuk menampilkan graph
@app.get("/graph", tags=["graph"])
async def visualize_graph(request_http: Request, token: str = Depends(verify_bearer_token)):
    file_path = "src/graph/graph-va-pmb-undiksha.png"
    if os.path.exists(file_path):
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"image/png\n###\nfilename={file_path}"
        })
        return FileResponse(file_path, media_type="image/png")
    else:
        raise HTTPException(status_code=404, detail="Tidak ditemukan file graph.")


# Enpoint untuk chat
@app.post("/chat", tags=["chat"])
async def chat_conversation(request: QuestionRequest, request_http: Request, token: str = Depends(verify_bearer_token)):
    question = request.question
    if not request.question:
        raise HTTPException(status_code=400, detail="Pertanyaan tidak boleh kosong.")
    try:
        _, answers = build_graph(question)
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 200,
            "success": True,
            "description": f"OK\n###\nQuestion:{question}\n###\nAnswer:{answers}"
        })
        return api_response(
            status_code=200,
            success=True,
            message="OK",
            data=[{
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "question": question,
                "answer": answers
            }]
        )
    except HTTPException as e:
        raise HTTPException(status_code=e.status_code, detail=f"{e.detail}")
    except Exception as e:
        log_activity({
            "id": generate_id(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "method": f"{request_http.method} {request_http.url.path}",
            "status_code": 500,
            "success": False,
            "description": f"Terjadi kesalahan yang tidak terduga. Pastikan LLM dan Embedder Service yang digunakan pada environment 'ollama' atau 'openai', lakukan proses embedding ulang, dan silahkan coba kembali. Jika masalah ini terus muncul, kemungkinan terdapat masalah pada LLM atau Embedder Service. {str(e)}"
        })
        return api_response(
            status_code=500,
            success=False,
            message=f"Terjadi kesalahan yang tidak terduga. Pastikan LLM dan Embedder Service yang digunakan pada environment 'ollama' atau 'openai', lakukan proses embedding ulang, dan silahkan coba kembali. Jika masalah ini terus muncul, kemungkinan terdapat masalah pada LLM atau Embedder Service. {str(e)}",
            data=None
        )


# Custom handler untuk 404 Not Found
@app.exception_handler(HTTP_404_NOT_FOUND)
async def not_found_handler(request: Request, exc: StarletteHTTPException):
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request.method} {request.url.path}",
        "status_code": 404,
        "success": False,
        "description": f"{exc.detail}"
    })
    return api_response(
        status_code=404,
        success=False,
        message=f"{exc.detail}",
        data=None
    )


# Custom handler untuk 405 Method Not Allowed
@app.exception_handler(HTTP_405_METHOD_NOT_ALLOWED)
async def method_not_allowed_handler(request: Request, exc: StarletteHTTPException):
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request.method} {request.url.path}",
        "status_code": 405,
        "success": False,
        "description": f"{exc.detail}"
    })
    return api_response(
        status_code=405,
        success=False,
        message=f"{exc.detail}",
        data=None
    )


# General handler untuk HTTP Exception lain
@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request: Request, exc: StarletteHTTPException):
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request.method} {request.url.path}",
        "status_code": exc.status_code,
        "success": False,
        "description": f"{exc.detail}"
    })
    return api_response(
        status_code=exc.status_code,
        success=False,
        message=f"{exc.detail}",
        data=None
    )


# General handler untuk validasi error
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    errors = exc.errors()
    error_messages = "; ".join([f"{err['loc']}: {err['msg']}" for err in errors])
    log_activity({
        "id": generate_id(),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": f"{request.method} {request.url.path}",
        "status_code": 422,
        "success": False,
        "description": f"{error_messages}"
    })
    return api_response(
        status_code=422,
        success=False,
        message=f"{error_messages}",
        data=None
    )


# RUNNING API
# uvicorn api.api:app --reload --port 3001